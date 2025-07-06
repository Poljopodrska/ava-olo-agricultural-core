#!/usr/bin/env python3
"""
AVA OLO Database Explorer API - Interactive database exploration
Connects to PostgreSQL database with Croatian agricultural data
"""

import logging
import io
import re
from datetime import datetime
from typing import Dict, Any, List, Optional, Union
from decimal import Decimal

from fastapi import FastAPI, HTTPException, Query, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import create_engine, text, inspect
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.exc import SQLAlchemyError
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import sys
import os

# Add parent directory to path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from config import DATABASE_URL, DB_POOL_SETTINGS

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Database connection
class DatabaseConnection:
    def __init__(self):
        self.engine = create_engine(DATABASE_URL, **DB_POOL_SETTINGS)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)
        self.inspector = inspect(self.engine)
    
    def get_session(self) -> Session:
        return self.SessionLocal()
    
    def get_tables(self) -> List[str]:
        """Get all table names"""
        return self.inspector.get_table_names()
    
    def get_columns(self, table_name: str) -> List[Dict]:
        """Get column information for a table"""
        try:
            columns = self.inspector.get_columns(table_name)
            return [
                {
                    "name": col["name"],
                    "type": str(col["type"]),
                    "nullable": col["nullable"],
                    "default": str(col["default"]) if col["default"] is not None else None,
                    "primary_key": col.get("primary_key", False)
                }
                for col in columns
            ]
        except Exception as e:
            logger.error(f"Error getting columns for {table_name}: {e}")
            return []
    
    def get_foreign_keys(self, table_name: str) -> List[Dict]:
        """Get foreign key information for a table"""
        try:
            fks = self.inspector.get_foreign_keys(table_name)
            return [
                {
                    "constrained_columns": fk["constrained_columns"],
                    "referred_table": fk["referred_table"],
                    "referred_columns": fk["referred_columns"]
                }
                for fk in fks
            ]
        except Exception as e:
            logger.error(f"Error getting foreign keys for {table_name}: {e}")
            return []

# Initialize database connection
db = DatabaseConnection()

# Pydantic models
class TableInfo(BaseModel):
    name: str
    row_count: int
    columns: List[Dict[str, Any]]
    foreign_keys: List[Dict[str, Any]]
    description: Optional[str] = None

class TableRelationship(BaseModel):
    from_table: str
    from_column: str
    to_table: str
    to_column: str
    relationship_type: str

class ColumnStats(BaseModel):
    name: str
    type: str
    nullable: bool
    null_count: int
    unique_count: int
    sample_values: List[str]
    min_value: Optional[str] = None
    max_value: Optional[str] = None

class TableData(BaseModel):
    table_name: str
    columns: List[str]
    rows: List[List[Any]]
    total_count: int
    page: int
    limit: int
    has_more: bool

# FastAPI app
app = FastAPI(
    title="AVA OLO Database Explorer API",
    description="Interactive database exploration for Croatian agricultural data",
    version="1.0.0"
)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Configure for production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Utility functions
def sanitize_table_name(table_name: str) -> str:
    """Sanitize table name to prevent SQL injection"""
    # Only allow alphanumeric characters, underscores
    if not re.match(r'^[a-zA-Z_][a-zA-Z0-9_]*$', table_name):
        raise HTTPException(status_code=400, detail="Invalid table name")
    return table_name

def serialize_value(value) -> str:
    """Convert database values to JSON-serializable strings"""
    if value is None:
        return ""
    elif isinstance(value, (datetime,)):
        return value.isoformat()
    elif isinstance(value, Decimal):
        return str(float(value))
    else:
        return str(value)

def build_where_clause(search: str, columns: List[str]) -> tuple:
    """Build WHERE clause for search functionality"""
    if not search:
        return "", {}
    
    # Search in text columns only
    text_columns = [col for col in columns if any(keyword in col.lower() 
                   for keyword in ['name', 'text', 'description', 'question', 'answer'])]
    
    if not text_columns:
        return "", {}
    
    conditions = []
    params = {}
    
    for i, col in enumerate(text_columns[:5]):  # Limit to 5 columns for performance
        param_name = f"search_param_{i}"
        conditions.append(f"LOWER({col}::text) LIKE LOWER(:{param_name})")
        params[param_name] = f"%{search}%"
    
    where_clause = f"WHERE {' OR '.join(conditions)}" if conditions else ""
    return where_clause, params

# API Endpoints

@app.get("/api/schema/tables", response_model=List[TableInfo])
async def get_tables():
    """Get all tables with basic information"""
    try:
        tables_info = []
        table_names = db.get_tables()
        
        with db.get_session() as session:
            for table_name in table_names:
                try:
                    # Get row count
                    row_count = session.execute(
                        text(f"SELECT COUNT(*) FROM {table_name}")
                    ).scalar()
                    
                    # Get columns and foreign keys
                    columns = db.get_columns(table_name)
                    foreign_keys = db.get_foreign_keys(table_name)
                    
                    # Table description based on name
                    descriptions = {
                        "ava_farmers": "Croatian farmers registered in the system",
                        "ava_fields": "Agricultural fields owned by farmers",
                        "ava_field_crops": "Crops planted in specific fields",
                        "ava_conversations": "Chat conversations with farmers",
                        "ava_crops": "Catalog of available crop types",
                        "farm_tasks": "Agricultural tasks and operations",
                        "system_health_log": "System health monitoring logs",
                        "llm_debug_log": "AI/LLM operation debugging logs",
                        "ava_weather": "Weather data for agricultural planning",
                        "ava_recommendations": "AI-generated farming recommendations"
                    }
                    
                    tables_info.append(TableInfo(
                        name=table_name,
                        row_count=row_count or 0,
                        columns=columns,
                        foreign_keys=foreign_keys,
                        description=descriptions.get(table_name)
                    ))
                    
                except Exception as table_error:
                    logger.warning(f"Error processing table {table_name}: {table_error}")
                    continue
        
        return tables_info
        
    except Exception as e:
        logger.error(f"Error in get_tables: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/schema/relationships", response_model=List[TableRelationship])
async def get_relationships():
    """Analyze foreign key relationships between tables"""
    try:
        relationships = []
        table_names = db.get_tables()
        
        for table_name in table_names:
            foreign_keys = db.get_foreign_keys(table_name)
            
            for fk in foreign_keys:
                for i, from_col in enumerate(fk["constrained_columns"]):
                    to_col = fk["referred_columns"][i] if i < len(fk["referred_columns"]) else "id"
                    
                    relationships.append(TableRelationship(
                        from_table=table_name,
                        from_column=from_col,
                        to_table=fk["referred_table"],
                        to_column=to_col,
                        relationship_type="foreign_key"
                    ))
        
        return relationships
        
    except Exception as e:
        logger.error(f"Error in get_relationships: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/data/{table_name}", response_model=TableData)
async def get_table_data(
    table_name: str,
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(50, ge=1, le=1000, description="Items per page"),
    search: str = Query("", description="Search term"),
    order_by: str = Query("", description="Column to order by"),
    order_dir: str = Query("desc", regex="^(asc|desc)$", description="Order direction")
):
    """Get paginated table data with search and filtering"""
    try:
        # Sanitize table name
        clean_table_name = sanitize_table_name(table_name)
        
        # Verify table exists
        if clean_table_name not in db.get_tables():
            raise HTTPException(status_code=404, detail="Table not found")
        
        with db.get_session() as session:
            # Get column information
            columns_info = db.get_columns(clean_table_name)
            column_names = [col["name"] for col in columns_info]
            
            # Build search clause
            where_clause, search_params = build_where_clause(search, column_names)
            
            # Build order clause
            order_clause = ""
            if order_by and order_by in column_names:
                order_clause = f"ORDER BY {order_by} {order_dir.upper()}"
            elif column_names:
                # Default ordering by first column
                order_clause = f"ORDER BY {column_names[0]} {order_dir.upper()}"
            
            # Get total count
            count_query = f"SELECT COUNT(*) FROM {clean_table_name} {where_clause}"
            total_count = session.execute(text(count_query), search_params).scalar()
            
            # Get paginated data
            offset = (page - 1) * limit
            data_query = f"""
                SELECT * FROM {clean_table_name} 
                {where_clause} 
                {order_clause} 
                LIMIT :limit OFFSET :offset
            """
            
            params = {**search_params, "limit": limit, "offset": offset}
            result = session.execute(text(data_query), params).fetchall()
            
            # Convert rows to serializable format
            rows = []
            for row in result:
                serialized_row = [serialize_value(value) for value in row]
                rows.append(serialized_row)
            
            has_more = (offset + limit) < total_count
            
            return TableData(
                table_name=clean_table_name,
                columns=column_names,
                rows=rows,
                total_count=total_count,
                page=page,
                limit=limit,
                has_more=has_more
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in get_table_data: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/data/{table_name}/columns", response_model=List[ColumnStats])
async def get_column_stats(table_name: str):
    """Get detailed column statistics for a table"""
    try:
        # Sanitize table name
        clean_table_name = sanitize_table_name(table_name)
        
        # Verify table exists
        if clean_table_name not in db.get_tables():
            raise HTTPException(status_code=404, detail="Table not found")
        
        with db.get_session() as session:
            columns_info = db.get_columns(clean_table_name)
            stats = []
            
            for col in columns_info:
                col_name = col["name"]
                
                try:
                    # Get null count
                    null_count = session.execute(
                        text(f"SELECT COUNT(*) FROM {clean_table_name} WHERE {col_name} IS NULL")
                    ).scalar()
                    
                    # Get unique count (limit for performance)
                    unique_count = session.execute(
                        text(f"SELECT COUNT(DISTINCT {col_name}) FROM {clean_table_name}")
                    ).scalar()
                    
                    # Get sample values
                    sample_result = session.execute(
                        text(f"SELECT DISTINCT {col_name} FROM {clean_table_name} WHERE {col_name} IS NOT NULL LIMIT 5")
                    ).fetchall()
                    sample_values = [serialize_value(row[0]) for row in sample_result]
                    
                    # Get min/max for appropriate column types
                    min_value = None
                    max_value = None
                    
                    if any(type_keyword in str(col["type"]).lower() 
                          for type_keyword in ['integer', 'numeric', 'decimal', 'timestamp', 'date']):
                        try:
                            min_result = session.execute(
                                text(f"SELECT MIN({col_name}) FROM {clean_table_name} WHERE {col_name} IS NOT NULL")
                            ).scalar()
                            max_result = session.execute(
                                text(f"SELECT MAX({col_name}) FROM {clean_table_name} WHERE {col_name} IS NOT NULL")
                            ).scalar()
                            
                            min_value = serialize_value(min_result)
                            max_value = serialize_value(max_result)
                        except:
                            pass  # Skip min/max for problematic columns
                    
                    stats.append(ColumnStats(
                        name=col_name,
                        type=str(col["type"]),
                        nullable=col["nullable"],
                        null_count=null_count or 0,
                        unique_count=unique_count or 0,
                        sample_values=sample_values,
                        min_value=min_value,
                        max_value=max_value
                    ))
                    
                except Exception as col_error:
                    logger.warning(f"Error analyzing column {col_name}: {col_error}")
                    # Add basic info even if stats fail
                    stats.append(ColumnStats(
                        name=col_name,
                        type=str(col["type"]),
                        nullable=col["nullable"],
                        null_count=0,
                        unique_count=0,
                        sample_values=[]
                    ))
            
            return stats
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in get_column_stats: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/data/{table_name}/export")
async def export_table_data(
    table_name: str,
    format: str = Query("excel", regex="^(excel|csv)$", description="Export format"),
    search: str = Query("", description="Search term"),
    limit: int = Query(10000, ge=1, le=50000, description="Maximum rows to export")
):
    """Export table data as Excel or CSV"""
    try:
        # Sanitize table name
        clean_table_name = sanitize_table_name(table_name)
        
        # Verify table exists
        if clean_table_name not in db.get_tables():
            raise HTTPException(status_code=404, detail="Table not found")
        
        with db.get_session() as session:
            # Get column information
            columns_info = db.get_columns(clean_table_name)
            column_names = [col["name"] for col in columns_info]
            
            # Build search clause
            where_clause, search_params = build_where_clause(search, column_names)
            
            # Get data
            query = f"SELECT * FROM {clean_table_name} {where_clause} LIMIT :limit"
            params = {**search_params, "limit": limit}
            result = session.execute(text(query), params).fetchall()
            
            # Convert to pandas DataFrame
            data = []
            for row in result:
                serialized_row = [serialize_value(value) for value in row]
                data.append(serialized_row)
            
            df = pd.DataFrame(data, columns=column_names)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            if format == "excel":
                # Create Excel file
                output = io.BytesIO()
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name=clean_table_name)
                    
                    # Format the Excel file
                    workbook = writer.book
                    worksheet = writer.sheets[clean_table_name]
                    
                    # Header formatting
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    
                    for col_num, column_title in enumerate(df.columns, 1):
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                        
                        # Auto-adjust column width
                        column_letter = openpyxl.utils.get_column_letter(col_num)
                        worksheet.column_dimensions[column_letter].width = min(
                            max(len(str(column_title)), 10), 50
                        )
                
                output.seek(0)
                
                headers = {
                    "Content-Disposition": f"attachment; filename={clean_table_name}_{timestamp}.xlsx",
                    "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                }
                
                return StreamingResponse(io.BytesIO(output.read()), headers=headers)
                
            else:  # CSV format
                output = io.StringIO()
                df.to_csv(output, index=False, encoding='utf-8')
                output.seek(0)
                
                headers = {
                    "Content-Disposition": f"attachment; filename={clean_table_name}_{timestamp}.csv",
                    "Content-Type": "text/csv; charset=utf-8"
                }
                
                return StreamingResponse(io.StringIO(output.getvalue()), headers=headers)
                
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in export_table_data: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/query/custom")
async def execute_custom_query(
    query: str = Query(..., description="SQL query to execute"),
    limit: int = Query(100, ge=1, le=1000, description="Result limit")
):
    """Execute custom SQL query (READ-ONLY)"""
    try:
        # Security: Only allow SELECT statements
        clean_query = query.strip().lower()
        if not clean_query.startswith('select'):
            raise HTTPException(status_code=400, detail="Only SELECT queries are allowed")
        
        # Additional security checks
        forbidden_keywords = ['insert', 'update', 'delete', 'drop', 'create', 'alter', 'truncate']
        if any(keyword in clean_query for keyword in forbidden_keywords):
            raise HTTPException(status_code=400, detail="Query contains forbidden keywords")
        
        with db.get_session() as session:
            # Add LIMIT if not present
            if 'limit' not in clean_query:
                query = f"{query} LIMIT {limit}"
            
            result = session.execute(text(query)).fetchall()
            
            if not result:
                return {"columns": [], "rows": [], "row_count": 0}
            
            # Get column names
            columns = list(result[0].keys()) if result else []
            
            # Convert rows
            rows = []
            for row in result:
                serialized_row = [serialize_value(value) for value in row]
                rows.append(serialized_row)
            
            return {
                "columns": columns,
                "rows": rows,
                "row_count": len(rows),
                "query": query
            }
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in execute_custom_query: {e}")
        raise HTTPException(status_code=500, detail=f"Query execution error: {str(e)}")

# Health check endpoint
@app.get("/health")
async def health_check():
    """Health check endpoint"""
    try:
        with db.get_session() as session:
            session.execute(text("SELECT 1"))
        
        return {
            "status": "healthy",
            "database": "connected",
            "tables": len(db.get_tables()),
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001, log_level="info")